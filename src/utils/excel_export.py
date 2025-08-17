import pandas as pd
import io
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def generate_excel(data_dict, filename_prefix, sheet_names=None, metadata=None):
    """
    Generate Excel file from dictionary of dataframes or a single dataframe.
    Returns (io.BytesIO, filename)
    """
    output = io.BytesIO()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{filename_prefix}_{timestamp}.xlsx"

    if isinstance(data_dict, pd.DataFrame):
        data_dict = {"Data": data_dict}

    if sheet_names and len(sheet_names) == len(data_dict):
        sheet_dict = {name: df for name, df in zip(sheet_names, data_dict.values())}
    else:
        sheet_dict = data_dict

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        if metadata:
            meta_items = [[k, v] for k, v in metadata.items()]
            meta_df = pd.DataFrame(meta_items, columns=['Propiedad', 'Valor'])
            meta_df.to_excel(writer, sheet_name='Metadatos', index=False)
            worksheet = writer.sheets['Metadatos']
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            for col_num, _ in enumerate(meta_df.columns.values, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
            for col_num, _ in enumerate(meta_df.columns, 1):
                worksheet.column_dimensions[get_column_letter(col_num)].width = 25

        for sheet_name, df in sheet_dict.items():
            clean_sheet_name = str(sheet_name).replace('/', '_').replace('\\', '_')
            clean_sheet_name = ''.join(c for c in clean_sheet_name if c.isalnum() or c in ['_', '-'])
            clean_sheet_name = clean_sheet_name[:31]

            df.to_excel(writer, sheet_name=clean_sheet_name, index=False)
            worksheet = writer.sheets[clean_sheet_name]
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            centered_alignment = Alignment(horizontal='center', vertical='center')
            for col_num, value in enumerate(df.columns.values, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = centered_alignment
            for col_num, column in enumerate(df.columns, 1):
                max_length = max([len(str(x)) for x in df[column].tolist()] + [len(str(column))]) + 2
                max_length = min(max_length, 40)
                worksheet.column_dimensions[get_column_letter(col_num)].width = max_length

    output.seek(0)
    return output, filename
